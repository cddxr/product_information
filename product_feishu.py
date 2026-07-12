from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import os
import json
import sqlite3
import hashlib
import re
import requests
from urllib.request import Request, urlopen
from urllib.parse import urlparse
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.core.driver_cache import DriverCacheManager
from dotenv import load_dotenv

load_dotenv()
from openpyxl.styles import PatternFill, Font, Alignment  #excel格式

options = Options()
options.add_argument("--window-size=1920,1080")
options.add_argument("--start-maximized")
options.add_argument("--disable-blink-features=AutomationControlled")

chrome_binary_path = os.getenv(
    "CHROME_BINARY_PATH",
    r"C:\Program Files\Google\Chrome\Application\chrome.exe"
).strip()

if not os.path.isfile(chrome_binary_path):
    raise FileNotFoundError(f"Chrome executable not found: {chrome_binary_path}")

options.binary_location = chrome_binary_path
driver_cache_dir = os.path.join(r"E:\amazon_data", "wdm-cache")
os.makedirs(driver_cache_dir, exist_ok=True)
service = Service(
    ChromeDriverManager(
        cache_manager=DriverCacheManager(root_dir=driver_cache_dir)
    ).install()
)
driver = webdriver.Chrome(service=service, options=options)
WAIT_LOAD = 5
WAIT_SCROLL = 3
WAIT_LONG = 20
WAIT_SHORT = 5
CONTINUE_RETRIES = 3

# def set_amazon_zipcode(driver, zipcode="10001"):
#     print(f"切换邮编到 {zipcode}")
#     driver.get("https://www.amazon.com/")
#     wait = WebDriverWait(driver, 60)
#
#     loc_btn = wait.until(
#         EC.element_to_be_clickable((By.ID, "nav-global-location-popover-link"))
#     )
#     loc_btn.click()
#
#     zip_input = wait.until(
#         EC.presence_of_element_located((By.ID, "GLUXZipUpdateInput"))
#     )
#     zip_input.clear()
#     zip_input.send_keys(zipcode)
#
#     driver.find_element(By.ID, "GLUXZipUpdate").click()
#
#     try:
#         time.sleep(1)
#         driver.find_element(By.NAME, "glowDoneButton").click()
#     except:
#         pass
#     time.sleep(2)

def scroll_and_capture(name, url, page_index):
    open_amazon_page(driver, url)

    screenshots = []
    last_height = driver.execute_script("return window.innerHeight")
    scroll_count = 0

    while True:
        screenshot_path = os.path.join(
            save_dir, f"{name}_page{page_index}_part{scroll_count+1}.png"
        )
        driver.save_screenshot(screenshot_path)
        screenshots.append(screenshot_path)

        driver.execute_script(f"window.scrollBy(0, {last_height});")
        scroll_count += 1
        time.sleep(WAIT_SCROLL)

        new_scroll_height = driver.execute_script(
            "return window.pageYOffset + window.innerHeight"
        )
        total_height = driver.execute_script("return document.body.scrollHeight")
        if new_scroll_height >= total_height:
            break

    return screenshots

def merge_images(image_paths, output_path):
    images = [Image.open(p) for p in image_paths]
    total_height = sum(img.height for img in images)
    max_width = max(img.width for img in images)

    merged_image = Image.new("RGB", (max_width, total_height))
    current_height = 0

    for img in images:
        merged_image.paste(img, (0, current_height))
        current_height += img.height

    merged_image.save(output_path)

    # 删除临时图
    for p in image_paths:
        os.remove(p)

def scroll_category_pages(category_name, first_page_url):
    url = first_page_url
    page_index = 1
    all_screenshots = []

    while url:
        screenshots = scroll_and_capture(category_name, url, page_index)
        all_screenshots.extend(screenshots)

        try:
            next_btn = driver.find_element(By.CSS_SELECTOR, "li.a-last a")
            url = next_btn.get_attribute("href")
            page_index += 1
        except:
            url = None

    return all_screenshots

def scroll_and_capture(name, page_index):
    """
    滚动当前页面并截图
    返回当前页截图列表
    """
    screenshots = []
    last_height = driver.execute_script("return window.innerHeight")
    scroll_count = 0

    while True:
        screenshot_path = os.path.join(
            save_dir, f"{name}_page{page_index}_part{scroll_count+1}.png"
        )
        driver.save_screenshot(screenshot_path)
        screenshots.append(screenshot_path)

        driver.execute_script(f"window.scrollBy(0, {last_height});")
        scroll_count += 1
        time.sleep(WAIT_SCROLL)

        new_scroll_height = driver.execute_script(
            "return window.pageYOffset + window.innerHeight"
        )
        total_height = driver.execute_script("return document.body.scrollHeight")
        if new_scroll_height >= total_height:
            break

    return screenshots


def merge_images(image_paths, output_path):
    images = [Image.open(p) for p in image_paths]
    total_height = sum(img.height for img in images)
    max_width = max(img.width for img in images)

    merged_image = Image.new("RGB", (max_width, total_height))
    current_height = 0

    for img in images:
        merged_image.paste(img, (0, current_height))
        current_height += img.height

    merged_image.save(output_path)

    # 删除临时图
    for p in image_paths:
        os.remove(p)


def scroll_category_pages(category_name, first_page_url, max_pages=10):

    url = first_page_url
    page_index = 1
    all_screenshots = []

    while url and page_index <= max_pages:
        print(f"     抓取 {category_name} 第 {page_index} 页...")
        open_amazon_page(driver, url)

        # 滚动截图当前页
        screenshots = scroll_and_capture(category_name, page_index)
        all_screenshots.extend(screenshots)

        # 尝试找到下一页
        try:
            next_btn = driver.find_element(By.CSS_SELECTOR, "li.a-last a")
            if next_btn.is_enabled():
                url = next_btn.get_attribute("href")
                page_index += 1
            else:
                break
        except:
            break

    return all_screenshots

#字典结构
file_structure ={
    "wsq.xlsx": {
        "蘑菇可可": {
            "products": {
                "VP": "https://www.amazon.com/dp/B0F8HXNY5N",
                "RYZE": "https://www.amazon.com/dp/B0DFRX59R2",
                "dreamcoco": "https://www.amazon.com/dp/B0FNR72SBX",
                "MoonBrew": "https://www.amazon.com/dp/B0F7YX5Y9W",
                "beam dream": "https://www.amazon.com/dp/B0CKY838PP",
                "Volocatilly": "https://www.amazon.com/dp/B0FF9KCX6S"
            }
        },
        "蛋白质咖啡": {
            "products": {
                "Venture Pal": "https://www.amazon.com/dp/B0FQCBMFYQ",
                "javvy": "https://www.amazon.com/dp/B0F9LS7ZKD",
                "chike": "https://www.amazon.com/dp/B00JBJR3HY",
                "FlavCity": "https://www.amazon.com/dp/B0BPK32N5S",
                "Bulletproof": "https://www.amazon.com/dp/B0FLFGGCVT",
            }
        },

        "叶绿素滴剂": {
            "products": {
                "Her Fantasy Box": "https://www.amazon.com/dp/B0DMQZ9J31?psc=1",
                "MaryRuth Organics": "https://www.amazon.com/dp/B08K3R68DZ?psc=1",
                "Double Wood Supplements ": "https://www.amazon.com/dp/B0BFJLWZV1?psc=1",
                "BENEVOLENT NOURISHMENT": "https://www.amazon.com/dp/B01AFI3IEG?psc=1",
                "Horbäach": "https://www.amazon.com/dp/B09BBPPNRD",
                "Now Food": "https://www.amazon.com/dp/B000MGWGOU?psc=1",
                "Alpha Flow": "https://www.amazon.com/dp/B097W72BHG?psc=1",
                "Nature's Way": "https://www.amazon.com/dp/B00016AGBQ?psc=1",
                "Juna ": "https://www.amazon.com/dp/B09SFFB7P2?psc=1",
                "Alpha Flow胶囊款": "https://www.amazon.com/dp/B096T7M8F2?psc=1",
            }
        },
        "艾草滴剂": {
            "products": {
                "Clean Nutraceuticals": "https://www.amazon.com/Wormwood-Turmeric-Berberine-Tincture-Clean/dp/B0D75PSYZV/ref=sr_1_2?crid=1BKTJJWTMZOHC&dib=eyJ2IjoiMSJ9.iyIJCH8hGbvht0TC4chA3Y_BWuQO_UM4LNUO0o1uWtx0hF2FDXVkOMM1xbmYb1wDY2WSA2n-9KFZvHLhDOXx646_lI866ZJNq7JirYYOuMPHpTGBeFLusF8xgIXS-Fgmqg1j8l03tEey7rgUsGEd-bCkuqt9qWSgfVo0PZ9rJu7y3uVVGbeC6sbilj5we679oAS3aZ2MNWOaaJkZbOr-hUXNAYXynwymCownMYHdWeVO6rYgykygRM5sKdGa5amAfzA5uMWVecDEG56k909fJW6g8xTXcCJTFGQpl5dDn6w.w5jDXl1NfhFSr8YhpOHnyeVpfSJeMmdA46xi8gfRbiI&dib_tag=se&keywords=wormwood%2Bblack%2Bwalnut%2Bclove%2Bparasite%2Bcleanse&qid=1769398501&rdc=1&sprefix=wormwood%2Bblack%2Bwalnut%2Bclove%2Bparasite%2Bcleanse%2Caps%2C366&sr=8-2&th=1",
                "VINATURA": "https://www.amazon.com/VINATURA-Wormwood-Black-Walnut-Capsules/dp/B0DGV8M1HH/ref=sr_1_1?crid=1BKTJJWTMZOHC&dib=eyJ2IjoiMSJ9.iyIJCH8hGbvht0TC4chA3Y_BWuQO_UM4LNUO0o1uWtx0hF2FDXVkOMM1xbmYb1wDY2WSA2n-9KFZvHLhDOXx646_lI866ZJNq7JirYYOuMPHpTGBeFLusF8xgIXS-Fgmqg1j8l03tEey7rgUsGEd-bCkuqt9qWSgfVo0PZ9rJu7y3uVVGbeC6sbilj5we679oAS3aZ2MNWOaaJkZbOr-hUXNAYXynwymCownMYHdWeVO6rYgykygRM5sKdGa5amAfzA5uMWVecDEG56k909fJW6g8xTXcCJTFGQpl5dDn6w.w5jDXl1NfhFSr8YhpOHnyeVpfSJeMmdA46xi8gfRbiI&dib_tag=se&keywords=wormwood+black+walnut+clove+parasite+cleanse&qid=1769398501&sprefix=wormwood+black+walnut+clove+parasite+cleanse%2Caps%2C366&sr=8-1",
                "Zahler": "https://www.amazon.com/Zahler-ParaGuard-Digestive-Intestinal-Certified/dp/B016LBZSDK/ref=sr_1_63?dib=eyJ2IjoiMSJ9.zIdyQxLYzNJ854D3Hn6lRjjhLRBROnoSG0iNOJFldZgDisRUF_yoDIZQfIGNXRHlTg-WJNtRM3m9wHWQ3972qFER9slHgoo7LLdEQbdinlA_iFWPIhPN0yUcr41w58i9IARc4677XK49zka1VJY-Q-OFAwsuJzsVGPqpjweKf_aBEdk5ixzWlQueVL8R4h0AwnYC4kTpSXBQqhY367kIhah8sqt8EG0U_hh8Hd21d9_4GnfVRDHg8I88URvQSSt9CX9QlGAHXCyslBiwoRn74ZMNOMxKSZ4Rtzwymgsjh7o.GIyvYwhAmAPtm_B5YiE3vKseM7oAzFq3TxCYMj5EdPk&dib_tag=se&keywords=wormwood%2Bblack%2Bwalnut%2Bclove%2Bparasite%2Bcleanse&qid=1769399566&sr=8-63&th=1",
                "Lukaree": "?crid=1BKTJJWTMZOHC&dib=eyJ2IjoiMSJ9.iyIJCH8hGbvht0TC4chA3Y_BWuQO_UM4LNUO0o1uWtx0hF2FDXVkOMM1xbmYb1wDY2WSA2n-9KFZvHLhDOXx646_lI866ZJNq7JirYYOuMPHpTGBeFLusF8xgIXS-Fgmqg1j8l03tEey7rgUsGEd-bCkuqt9qWSgfVo0PZ9rJu7y3uVVGbeC6sbilj5we679oAS3aZ2MNWOaaJkZbOr-hUXNAYXynwymCownMYHdWeVO6rYgykygRM5sKdGa5amAfzA5uMWVecDEG56k909fJW6g8xTXcCJTFGQpl5dDn6w.w5jDXl1NfhFSr8YhpOHnyeVpfSJeMmdA46xi8gfRbiI&dib_tag=se&keywords=wormwood%2Bblack%2Bwalnut%2Bclove%2Bparasite%2Bcleanse&qid=1769398501&sprefix=wormwood%2Bblack%2Bwalnut%2Bclove%2Bparasite%2Bcleanse%2Caps%2C366&sr=8-4&th=1",
                "NOW Foods": "https://www.amazon.com/Foods-Walnut-Wormwood-Complex-2-Ounce/dp/B000S92RNQ/ref=sr_1_9?crid=1BKTJJWTMZOHC&dib=eyJ2IjoiMSJ9.iyIJCH8hGbvht0TC4chA3Y_BWuQO_UM4LNUO0o1uWtx0hF2FDXVkOMM1xbmYb1wDY2WSA2n-9KFZvHLhDOXx646_lI866ZJNq7JirYYOuMPHpTGBeFLusF8xgIXS-Fgmqg1j8l03tEey7rgUsGEd-bCkuqt9qWSgfVo0PZ9rJu7y3uVVGbeC6sbilj5we679oAS3aZ2MNWOaaJkZbOr-hUXNAYXynwymCownMYHdWeVO6rYgykygRM5sKdGa5amAfzA5uMWVecDEG56k909fJW6g8xTXcCJTFGQpl5dDn6w.w5jDXl1NfhFSr8YhpOHnyeVpfSJeMmdA46xi8gfRbiI&dib_tag=se&keywords=wormwood+black+walnut+clove+parasite+cleanse&qid=1769398501&sprefix=wormwood+black+walnut+clove+parasite+cleanse%2Caps%2C366&sr=8-9"
            }
        },
        "类目表格": {
            "categories": {
                "Protein Coffee": "https://www.amazon.com/gp/bestsellers/grocery/14808786011",
                "cold brew": "https://www.amazon.com/gp/bestsellers/grocery/14808786011/ref=pd_zg_hrsr_grocery",
                "叶绿素市场":"https://www.amazon.com/gp/bestsellers/hpc/6966318011/ref=pd_zg_hrsr_hpc",
                "草本市场":"https://www.amazon.com/gp/bestsellers/hpc/3764461/ref=pd_zg_hrsr_hpc",

            }
        }
    },

    "猫镇静软咀嚼、湿巾们.xlsx": {
        "蘑菇粉": {
            "products": {
                "Petsmont": "https://www.amazon.com/dp/B08D6VWVMS?psc=1",
                "Pets Are Kids Too": "https://www.amazon.com/dp/B0DZ4VQ3K5?psc=1",
                "VetSmart Formulas": "https://www.amazon.com/dp/B07FWPPGWT?psc=1",
                "Fera Pets": "https://www.amazon.com/dp/B08ZHZ6CZ4?psc=1",
                "WONDER PAWS": "https://www.amazon.com/dp/B0DVZQS2XR?psc=1",
                "Ninsiec": "https://www.amazon.com/Ninsiec-Mushroom-Supplement-Powder-Dogs/dp/B0DSSHKFN7"
            }
        },
    },
        "tx.xlsx": {
                "vp咖啡": {
                    "products": {
                        "Venture Pal": "https://www.amazon.com/dp/B0DDCJFFBM",
                        "RYZE SUPERFOODS": "https://www.amazon.com/dp/B0DJWV4BQ1",
                        "Everyday Dose": "https://www.amazon.com/dp/B0B53Y7Z7D",
                        "Four Sigmatic": "https://www.amazon.com/dp/B0756D1D39",
                        "MAX FIT": "https://www.amazon.com/dp/B0C8WGD562",
                        "Conscious Beyond": "https://www.amazon.com/dp/B0B628D7ZN",
                        "Dandy Blend": "https://www.amazon.com/dp/B000SMN0DO",
                        "Slim Coffee": "https://www.amazon.com/dp/B0DFBMVX7T",
                        "GRGTP":"https://www.amazon.com/GRGTP-Mushroom-Coffee-Adaptogens-Jitter-Free/dp/B0FH2GPBQH/ref=sr_1_23?",
                        "Bcuelov":"https://www.amazon.com/Mushroom-Adaptogenic-Mushrooms-Cordyceps-Shiitake/dp/B0FLPQVN6X/ref=sr_1_11?",
                        "YEGE":"https://www.amazon.com/Mushroom-Instant-Cordyceps-Positive-Servings/dp/B0DJSKQRPM/ref=sr_1_4?",
                        "Bunkell":"https://www.amazon.com/Bunkell-Mushroom-Mushrooms-Cordyceps-Digestion/dp/B0DHP915W6/ref=sr_1_1?th=1",
                    }
                },

            "解酒": {
                "products": {
                    "Venture Pal": "https://www.amazon.com/dp/B0DGPZ81BC?th=1",
                    "Waterboy": "https://www.amazon.com/dp/B0FKL2M4GH?th=1",
                    "Liquid I.V.": "https://www.amazon.com/dp/B0CSQWZPBX?th=1",
                    "Nectar": "https://www.amazon.com/dp/B09BD8GZ8L?th=1",
                    "Ultima Replenisher": "https://www.amazon.com/dp/B08XQZX9K3?th=1",
                }
            },

                "成分虚标": {
                    "products": {
                        "fresh cap": "https://www.amazon.com/dp/B0G1Z7XH1S",
                        "suiyilary": "https://www.amazon.com/dp/B0DN1X25D8",
                        "wellness_1": "https://www.amazon.com/dp/B0CLT4SLGD",
                        "wellness_2": "https://www.amazon.com/dp/B0DCZG76K4",
                        "toniiq": "https://www.amazon.com/dp/B0D41N4J1Y"
                    }
                },
                "类目列表": {
                    "categories": {
                        "instant coffee": "https://www.amazon.com/gp/bestsellers/grocery/2251594011/ref=pd_zg_hrsr_grocery",
                        "运动电解质市场":"https://www.amazon.com/gp/bestsellers/hpc/6973694011/ref=pd_zg_hrsr_hpc"
                    }
                }
            },

        "lyj.xlsx": {
            "功能电解质": {
                "products": {
                    "Hunhe": "https://www.amazon.com/Venture-Pal-Sugar-Free-Electrolyte-Packets/dp/B0DN6T1QGD/ref=cm_cr_arp_d_product_top?ie=UTF8&th=1",
                    "Xigua": "https://www.amazon.com/Venture-Pal-Sugar-Free-Electrolyte-Packets/dp/B0DN6TX81K/ref=cm_cr_arp_d_product_top?ie=UTF8&th=1",
                    "Ningmeng": "https://www.amazon.com/Venture-Pal-Sugar-Free-Electrolyte-Packets/dp/B0DN6QV16L/ref=cm_cr_arp_d_product_top?ie=UTF8&th=1",
                    "Meiguo": "https://www.amazon.com/Venture-Pal-Sugar-Free-Electrolyte-Packets/dp/B0DN6QTZNG/ref=cm_cr_arp_d_product_top?ie=UTF8&th=1",
                    "Taozi": "https://www.amazon.com/Venture-Pal-Sugar-Free-Electrolyte-Packets/dp/B0DN6QG9YN/ref=cm_cr_arp_d_product_top?ie=UTF8&th=1",
                    "LMNT": "https://www.amazon.com/LMNT-Zero-Sugar-Electrolytes/dp/B0FTGJGPTM/?th=1",
                    "REDMOND": "https://www.amazon.com/REDMOND-Re-Lyte-Electrolyte-Strawberry-Lemonade/dp/B097QF7HVH/",
                    "Santa Cruz": "https://www.amazon.com/Santa-Cruz-Paleo-Electrolyte-Sugar-Free/dp/B0F1CDC3KT/",
                    "SALTT": "https://www.amazon.com/SALTT-Electrolytes-Powder-Unflavored-Drink/dp/B0CQRPX3HC/",
                    "Totaria新品":"https://www.amazon.com/Totaria-Electrolyte-Hydration-Himalayan-Electrolytes/dp/B0GX8G9L3G"
                }
            },

            "普通电解质": {
                "products": {
                    "Waterboy": "https://www.amazon.com/dp/B0FKL2M4GH?th=1",
                    "Liquid I.V.": "https://www.amazon.com/Sugar-Free-Lemon-Lime-14/dp/B0BQ4G7LY8/",
                    "Nectar": "https://www.amazon.com/dp/B09BD8GZ8L?th=1",
                    "Ultima Replenisher": "https://www.amazon.com/Ultima-Replenisher-Daily-Electrolyte-Powder/dp/B0FMGNZYNL", #update by wsq 26.6.23
                    "Cure": "https://www.amazon.com/Cure-Hydration-Natural-Electrolyte-Friendly/dp/B07RWF8L2V/",
                    "DripDrop": "https://www.amazon.com/DripDrop-ORS-Electrolyte-Dehydration-Watermelon/dp/B08TZD2R7L/",
                    "liquid iv 有糖":"https://www.amazon.com/Liquid-I-V-Multiplier-Electrolyte-Supplement/dp/B01IT9NLHW"
                }
            },
            "益智电解质": {
                "products": {
                    "Venture Pal": "https://www.amazon.com/gp/product/B0FKN8PWNJ?th=1",
                    "IQMIX": "https://www.amazon.com/dp/B09MQ54VPC?psc=1",
                    "Ultima": "https://www.amazon.com/dp/B0FL4SJS9W"
                }
            },

            "分类": {
                "categories": {
                    "Sports Nutrition Electrolyte Drinks": "https://www.amazon.com/gp/bestsellers/hpc/6973694011/ref=pd_zg_hrsr_hpc",
                    "钠矿物质补充剂":"https://www.amazon.com/gp/bestsellers/hpc/3774491"
                }
        }
    },
    "zlq.xlsx": {
        "高钠": {
            "products": {
                "Venture Pal": "https://www.amazon.com/gp/product/B0DN6T1QGD",
                "LMNT": "https://www.amazon.com/LMNT-Zero-Sugar-Electrolytes/dp/B0FTGJGPTM/?th=1",
                "REDMOND": "https://www.amazon.com/REDMOND-Re-Lyte-Electrolyte-Strawberry-Lemonade/dp/B097QF7HVH/",
                "Santa Cruz": "https://www.amazon.com/Santa-Cruz-Paleo-Electrolyte-Sugar-Free/dp/B0F1CDC3KT/",
                "SALTT": "https://www.amazon.com/SALTT-Electrolytes-Powder-Unflavored-Drink/dp/B0CQRPX3HC/"
            }
        },
        "Vp": {
            "products": {
                "草莓柠檬": "https://www.amazon.com/dp/B0FKN6RCLB/ref=twister_B0GQ327HZQ?th=1",
                "蓝树莓": "https://www.amazon.com/dp/B0FKN3CHKP/ref=twister_B0GQ327HZQ?th=1",
                "菠萝": "https://www.amazon.com/dp/B0FKN5CN6W/ref=twister_B0GQ327HZQ?th=1",
                "樱桃": "https://www.amazon.com/dp/B0FKN5Q812/ref=twister_B0GQ327HZQ?th=1",
                "冰棒": "https://www.amazon.com/dp/B0FKN9KSV1/ref=twister_B0GQ327HZQ?th=1",
                "巧克力": "https://www.amazon.com/dp/B0FKN6TPDF/ref=twister_B0GQ327HZQ?th=1",
                "水果潘趣": "https://www.amazon.com/dp/B0FKN6MF2Y/ref=twister_B0GQ327HZQ?th=1",
                "葡萄": "https://www.amazon.com/dp/B0FKN577JS/ref=twister_B0GQ327HZQ?th=1",
                "VP016-3": "https://www.amazon.com/Venture-Pal-Electrolytes-Hydration-Electrolyte/dp/B0FKN3ZTBJ/ref",
                "VP016-4": "https://www.amazon.com/Venture-Pal-Electrolytes-Hydration-Electrolyte/dp/B0FVLNKFJL/ref=zg_bs_g_6973694011_d_sccl_81/147-6123590-1087668?th=1"
            }
        },

        "铁胶囊": {
            "products": {
                "Megafood": "https://www.amazon.com/dp/B000F4ZRCC?psc=1",
                "THORNE": "https://www.amazon.com/dp/B0797PJQMT?psc=1",
                "Nutra Harmony": "https://www.amazon.com/dp/B0D47FBQP6?psc=1",
            }
        },
        "分类": {
            "categories": {
                "Sports Nutrition Electrolyte Drinks": "https://www.amazon.com/gp/bestsellers/hpc/6973694011/ref=pd_zg_hrsr_hpc",
                "Sports Nutrition Endurance & Energy Powders": "https://www.amazon.com/gp/bestsellers/hpc/6973677011/ref=pd_zg_hrsr_hpc"
            }
        }
    },
    "lyw.xlsx": {
        "高钠": {
            "products": {
                "Venture Pal": "https://www.amazon.com/gp/product/B0DN6T1QGD",
                "LMNT": "https://www.amazon.com/LMNT-Zero-Sugar-Electrolytes/dp/B0FTGJGPTM/?th=1",
                "REDMOND": "https://www.amazon.com/REDMOND-Re-Lyte-Electrolyte-Strawberry-Lemonade/dp/B097QF7HVH/",
                "Totaria": "https://www.amazon.com/Totaria-Electrolyte-Hydration-Himalayan-Electrolytes/dp/B0GX8G9L3G/ref=",
                "SALTT": "https://www.amazon.com/SALTT-Electrolytes-Powder-Unflavored-Drink/dp/B0CQRPX3HC/"
            }
        },
        "Vp": {
            "products": {
                "草莓柠檬": "https://www.amazon.com/dp/B0FKN6RCLB/ref=twister_B0GQ327HZQ?th=1",
                "蓝树莓": "https://www.amazon.com/dp/B0FKN3CHKP/ref=twister_B0GQ327HZQ?th=1",
                "菠萝": "https://www.amazon.com/dp/B0FKN5CN6W/ref=twister_B0GQ327HZQ?th=1",
                "樱桃": "https://www.amazon.com/dp/B0FKN5Q812/ref=twister_B0GQ327HZQ?th=1",
                "冰棒": "https://www.amazon.com/dp/B0FKN9KSV1/ref=twister_B0GQ327HZQ?th=1",
                "巧克力": "https://www.amazon.com/dp/B0FKN6TPDF/ref=twister_B0GQ327HZQ?th=1",
                "水果潘趣": "https://www.amazon.com/dp/B0FKN6MF2Y/ref=twister_B0GQ327HZQ?th=1",
                "葡萄": "https://www.amazon.com/dp/B0FKN577JS/ref=twister_B0GQ327HZQ?th=1",
                "VP016-3": "https://www.amazon.com/Venture-Pal-Electrolytes-Hydration-Electrolyte/dp/B0FKN3ZTBJ/ref",
                "VP016-4": "https://www.amazon.com/Venture-Pal-Electrolytes-Hydration-Electrolyte/dp/B0FVLNKFJL/ref=zg_bs_g_6973694011_d_sccl_81/147-6123590-1087668?th=1"
            }
        },
        "解酒": {
            "products": {
                "Venture Pal": "https://www.amazon.com/dp/B0DGPZ81BC?th=1",
                "Waterboy": "https://www.amazon.com/dp/B0FKL2M4GH?th=1",
                "Liquid I.V.": "https://www.amazon.com/Sugar-Free-Lemon-Lime-14/dp/B0BQ4G7LY8/",
                "Nectar": "https://www.amazon.com/dp/B09BD8GZ8L?th=1",
                "Ultima Replenisher": "https://www.amazon.com/dp/B08XQZX9K3?th=1",
            }
        },

            "分类": {
                "categories": {
                    "Sports Nutrition Electrolyte Drinks": "https://www.amazon.com/gp/bestsellers/hpc/6973694011/ref=pd_zg_hrsr_hpc",
                    "Sports Nutrition Endurance & Energy Powders": "https://www.amazon.com/gp/bestsellers/hpc/6973677011/ref=pd_zg_hrsr_hpc",
                    "铁胶囊":"https://www.amazon.com/gp/bestsellers/hpc/3774401/ref=pd_zg_hrsr_hpc",
                }
            }
        }
    }

save_dir = r"E:\amazon_data\product_information"
os.makedirs(save_dir, exist_ok=True)
db_path = os.path.join(save_dir, "amazon_product_snapshots.sqlite3")


SNAPSHOT_EXTRA_COLUMNS = {
    "price": "TEXT",
    "stars": "TEXT",
    "rating": "TEXT",
    "total_reviews": "TEXT",
    "flavor_count": "TEXT",
    "size_count": "TEXT",
}


COMPARE_FIELDS = [
    ("sheet_name", "类目"),
    ("brand", "品牌名"),
    ("url", "链接"),
    ("title", "标题"),
    ("price", "价格"),
    ("discount", "折扣/Deal"),
    ("coupons", "促销方式"),
    ("flavor_count", "口味数量"),
    ("size_count", "规格数量"),
    ("stars", "星级占比"),
    ("rating", "评分"),
    ("main_image_url", "主图链接"),
    ("sub_image_urls", "副图链接"),
]


def ensure_columns(conn, table_name, columns):
    existing = {
        row[1]
        for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    }
    for column_name, column_type in columns.items():
        if column_name not in existing:
            conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")


def init_database(path):
    conn = sqlite3.connect(path)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS product_latest (
            product_key TEXT PRIMARY KEY,
            file_name TEXT,
            sheet_name TEXT,
            brand TEXT,
            url TEXT,
            title TEXT,
            discount TEXT,
            coupons TEXT,
            promotion_text TEXT,
            main_image_url TEXT,
            sub_image_urls TEXT,
            main_image_ocr TEXT,
            sub_image_ocr TEXT,
            price TEXT,
            stars TEXT,
            rating TEXT,
            total_reviews TEXT,
            flavor_count TEXT,
            size_count TEXT,
            captured_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS product_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_key TEXT,
            file_name TEXT,
            sheet_name TEXT,
            brand TEXT,
            url TEXT,
            title TEXT,
            discount TEXT,
            coupons TEXT,
            promotion_text TEXT,
            main_image_url TEXT,
            sub_image_urls TEXT,
            main_image_path TEXT,
            sub_image_paths TEXT,
            main_image_ocr TEXT,
            sub_image_ocr TEXT,
            price TEXT,
            stars TEXT,
            rating TEXT,
            total_reviews TEXT,
            flavor_count TEXT,
            size_count TEXT,
            remark TEXT,
            captured_at TEXT
        )
        """
    )
    ensure_columns(conn, "product_latest", SNAPSHOT_EXTRA_COLUMNS)
    ensure_columns(conn, "product_snapshots", SNAPSHOT_EXTRA_COLUMNS)
    conn.commit()
    return conn


def product_key(file_name, sheet_name, brand, url):
    raw = f"{file_name}|{sheet_name}|{brand}|{url.strip()}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def normalize_text(value):
    return " ".join((value or "").split())


def normalize_list(values):
    return [normalize_text(v) for v in values if normalize_text(v)]


def promotion_text(discount, coupons):
    parts = []
    if normalize_text(discount):
        parts.append(normalize_text(discount))
    parts.extend(normalize_list(coupons))
    return "; ".join(parts) if parts else "无"


def format_promotion_item(text):
    text = normalize_text(text)
    if not text or text == "无":
        return ""

    subscribe_match = re.search(r"Save\s+(\d+)%.*Subscribe", text, re.IGNORECASE)
    if subscribe_match:
        return f"订阅省{subscribe_match.group(1)}%"

    coupon_match = re.search(r"(?:Apply\s+)?(\d+)%\s+coupon", text, re.IGNORECASE)
    if coupon_match:
        return f"{coupon_match.group(1)}%coupon"

    code_match = re.search(r"Save\s+(\d+)%\s+on\s+(\d+)\s+select item", text, re.IGNORECASE)
    if code_match:
        discount_percent, item_count = code_match.groups()
        return f"满{item_count}减{discount_percent}%code"

    qualify_match = re.search(r"Up to\s+(\d+)%\s+off\s+if you qualify", text, re.IGNORECASE)
    if qualify_match:
        return f"符合条件最高省{qualify_match.group(1)}%"

    return text.replace(" Terms", "").replace(" Shop items", "").strip()


def format_promotions(discount, coupons):
    formatted = []
    for item in [discount] + list(coupons or []):
        formatted_item = format_promotion_item(item)
        if formatted_item and formatted_item not in formatted:
            formatted.append(formatted_item)
    return "；".join(formatted) if formatted else "无"


def load_latest_snapshot(conn, key):
    conn.row_factory = sqlite3.Row
    return conn.execute(
        "SELECT * FROM product_latest WHERE product_key = ?",
        (key,),
    ).fetchone()


def compare_with_previous(previous, current):
    if previous is None:
        return "首次抓取"

    changes = []
    for field_name, label in COMPARE_FIELDS:
        previous_value = previous[field_name] if field_name in previous.keys() else None
        if previous_value is None:
            continue
        if normalize_text(previous_value) != normalize_text(current.get(field_name)):
            changes.append(f"{label}发生变化")

    return "；".join(changes) if changes else "无变化"


def save_successful_snapshot(conn, current):
    conn.execute(
        """
        INSERT INTO product_snapshots (
            product_key, file_name, sheet_name, brand, url, title, discount, coupons,
            promotion_text, main_image_url, sub_image_urls, main_image_path,
            sub_image_paths, main_image_ocr, sub_image_ocr, price, stars, rating,
            total_reviews, flavor_count, size_count, remark, captured_at
        )
        VALUES (
            :product_key, :file_name, :sheet_name, :brand, :url, :title, :discount,
            :coupons, :promotion_text, :main_image_url, :sub_image_urls,
            :main_image_path, :sub_image_paths, :main_image_ocr, :sub_image_ocr,
            :price, :stars, :rating, :total_reviews, :flavor_count, :size_count,
            :remark, :captured_at
        )
        """,
        current,
    )
    conn.execute(
        """
        INSERT INTO product_latest (
            product_key, file_name, sheet_name, brand, url, title, discount, coupons,
            promotion_text, main_image_url, sub_image_urls, main_image_ocr,
            sub_image_ocr, price, stars, rating, total_reviews, flavor_count,
            size_count, captured_at
        )
        VALUES (
            :product_key, :file_name, :sheet_name, :brand, :url, :title, :discount,
            :coupons, :promotion_text, :main_image_url, :sub_image_urls,
            :main_image_ocr, :sub_image_ocr, :price, :stars, :rating,
            :total_reviews, :flavor_count, :size_count, :captured_at
        )
        ON CONFLICT(product_key) DO UPDATE SET
            file_name = excluded.file_name,
            sheet_name = excluded.sheet_name,
            brand = excluded.brand,
            url = excluded.url,
            title = excluded.title,
            discount = excluded.discount,
            coupons = excluded.coupons,
            promotion_text = excluded.promotion_text,
            main_image_url = excluded.main_image_url,
            sub_image_urls = excluded.sub_image_urls,
            main_image_ocr = excluded.main_image_ocr,
            sub_image_ocr = excluded.sub_image_ocr,
            price = excluded.price,
            stars = excluded.stars,
            rating = excluded.rating,
            total_reviews = excluded.total_reviews,
            flavor_count = excluded.flavor_count,
            size_count = excluded.size_count,
            captured_at = excluded.captured_at
        """,
        current,
    )
    conn.commit()


def first_dynamic_image_url(element):
    for attr in ("data-old-hires", "src"):
        value = element.get_attribute(attr)
        if value:
            return value

    dynamic_images = element.get_attribute("data-a-dynamic-image")
    if dynamic_images:
        try:
            image_map = json.loads(dynamic_images)
            urls = list(image_map.keys())
            if urls:
                return urls[0]
        except json.JSONDecodeError:
            pass
    return ""


def add_image_url(image_urls, url):
    url = (url or "").strip()
    if not url or url.startswith("data:"):
        return
    if url not in image_urls:
        image_urls.append(url)


def best_image_url_from_item(item):
    if not isinstance(item, dict):
        return ""
    for key in ("hiRes", "large", "mainUrl", "variantUrl", "thumb"):
        value = item.get(key)
        if value:
            return value
    return ""


def image_urls_from_landing_image(driver):
    image_urls = []
    try:
        landing_image = driver.find_element(By.ID, "landingImage")
    except:
        return image_urls

    dynamic_images = landing_image.get_attribute("data-a-dynamic-image")
    if dynamic_images:
        try:
            image_map = json.loads(dynamic_images)
            for image_url in image_map.keys():
                add_image_url(image_urls, image_url)
        except json.JSONDecodeError:
            pass

    add_image_url(image_urls, landing_image.get_attribute("data-old-hires"))
    add_image_url(image_urls, landing_image.get_attribute("src"))
    return image_urls


def image_urls_from_amazon_scripts(driver):
    script = """
        const urls = [];
        function add(value) {
            if (value && typeof value === 'string' && !urls.includes(value)) {
                urls.push(value);
            }
        }
        function readItem(item) {
            if (!item || typeof item !== 'object') return;
            add(item.hiRes);
            add(item.large);
            add(item.mainUrl);
            add(item.variantUrl);
            add(item.thumb);
        }
        const blocks = [
            window.ImageBlockATF,
            window.ImageBlockBTF,
            window.P && window.P._namespace && window.P._namespace('ImageBlockATF')
        ];
        for (const block of blocks) {
            const initial = block && block.colorImages && block.colorImages.initial;
            if (Array.isArray(initial)) {
                initial.forEach(readItem);
            }
        }
        return urls;
    """
    image_urls = []
    try:
        for image_url in driver.execute_script(script) or []:
            add_image_url(image_urls, image_url)
    except:
        pass

    page_source = driver.page_source
    for pattern in (
        r'"hiRes"\s*:\s*"([^"]+)"',
        r'"large"\s*:\s*"([^"]+)"',
        r'"mainUrl"\s*:\s*"([^"]+)"',
    ):
        for match in re.findall(pattern, page_source):
            add_image_url(image_urls, match.replace("\\/", "/"))

    return image_urls


def image_urls_from_thumbnail_clicks(driver):
    image_urls = []
    thumbnails = driver.find_elements(By.CSS_SELECTOR, "#altImages li.imageThumbnail, #altImages li.a-spacing-small")
    for thumb in thumbnails:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", thumb)
            thumb.click()
            time.sleep(0.5)
            for image_url in image_urls_from_landing_image(driver):
                add_image_url(image_urls, image_url)
        except:
            continue
    return image_urls


def get_product_image_urls(driver):
    image_urls = []
    for extractor in (
        image_urls_from_landing_image,
        image_urls_from_amazon_scripts,
        image_urls_from_thumbnail_clicks,
    ):
        for image_url in extractor(driver):
            add_image_url(image_urls, image_url)

    return image_urls


def safe_file_name(value):
    keep = []
    for ch in value:
        keep.append(ch if ch.isalnum() or ch in ("-", "_") else "_")
    return "".join(keep).strip("_")[:80] or "image"


def download_image(url, output_dir, prefix, index):
    if not url:
        return ""
    os.makedirs(output_dir, exist_ok=True)
    suffix = os.path.splitext(urlparse(url).path)[1].split("?")[0]
    if suffix.lower() not in (".jpg", ".jpeg", ".png", ".webp"):
        suffix = ".jpg"
    output_path = os.path.join(output_dir, f"{safe_file_name(prefix)}_{index}{suffix}")
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=30) as response:
            with open(output_path, "wb") as f:
                f.write(response.read())
        return output_path
    except Exception as e:
        print(f"图片下载失败: {url} | {e}")
        return ""


def download_product_images(image_urls, output_dir, prefix):
    paths = []
    for index, image_url in enumerate(image_urls, start=1):
        path = download_image(image_url, output_dir, prefix, index)
        if path:
            paths.append(path)
    return paths


def image_ocr_text(image_path):
    if not image_path:
        return ""
    try:
        import pytesseract
        return normalize_text(pytesseract.image_to_string(Image.open(image_path)))
    except ImportError:
        return ""
    except Exception as e:
        print(f"OCR识别失败: {image_path} | {e}")
        return ""


def handle_amazon_continue(driver, retries=CONTINUE_RETRIES):
    continue_xpath = (
        "//button[contains(translate(normalize-space(.), "
        "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'continue')]"
        " | //input[contains(translate(@value, "
        "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'continue')]"
        " | //a[contains(translate(normalize-space(.), "
        "'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'continue')]"
        " | //input[@name='continue']"
        " | //button[@name='continue']"
    )

    def visible_continue_buttons():
        buttons = []
        for element in driver.find_elements(By.XPATH, continue_xpath):
            try:
                if element.is_displayed() and element.is_enabled():
                    buttons.append(element)
            except Exception:
                continue
        return buttons

    def is_continue_page():
        if visible_continue_buttons():
            return True
        try:
            page_text = driver.find_element(By.TAG_NAME, "body").text.lower()
        except Exception:
            page_text = ""
        return (
            "continue shopping" in page_text
            or "click the button below to continue" in page_text
        )

    for attempt in range(1, retries + 1):
        try:
            if not is_continue_page():
                return True

            buttons = visible_continue_buttons()
            if buttons:
                button = buttons[0]
                driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center'});", button
                )
                time.sleep(0.5)
                try:
                    button.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", button)
                print(f"检测到 Amazon Continue 页面，第 {attempt} 次自动点击")
                time.sleep(WAIT_LOAD)

            if not is_continue_page():
                return True
        except Exception as e:
            print(f"Continue 页面处理失败，第 {attempt} 次: {e}")

        try:
            driver.refresh()
            time.sleep(WAIT_LOAD)
        except:
            pass

    return False


def open_amazon_page(driver, url):
    for attempt in range(1, CONTINUE_RETRIES + 1):
        driver.get(url)
        time.sleep(WAIT_LOAD)
        if handle_amazon_continue(driver):
            return
        print(f"Amazon Continue 页面未通过，重新打开页面，第 {attempt} 次")
    raise RuntimeError("Amazon Continue 页面多次处理失败")


def wait_title(driver):
    return WebDriverWait(driver, WAIT_LONG).until(
        EC.presence_of_element_located((By.ID, "productTitle"))
    ).text.strip()

def get_price(driver):
    try:
        whole = driver.find_element(By.CSS_SELECTOR, "span.a-price-whole").text
        frac = driver.find_element(By.CSS_SELECTOR, "span.a-price-fraction").text
        symbol = driver.find_element(By.CSS_SELECTOR, "span.a-price-symbol").text
        return f"{symbol}{whole}.{frac}"
    except:
        return "未找到"

def get_coupon(driver):
    coupons = []
    try:
        elements = WebDriverWait(driver, WAIT_SHORT).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "span.couponLabelText"))
        )
        for el in elements:
            text = el.text.strip()
            if text and text not in coupons:
                coupons.append(text)
    except:
        pass
    try:
        green_labels = driver.find_elements(By.CSS_SELECTOR, "label[id^='greenBadge']")
        for green in green_labels:
            parent_text = green.find_element(By.XPATH, "..").text.strip()
            if parent_text and parent_text not in coupons:
                coupons.append(parent_text)
    except:
        pass
    return coupons

def get_discount(driver):
    try:
        el = driver.find_element(By.CSS_SELECTOR, "span.savingPriceOverride, span.dealBadge")
        return el.text.strip()
    except:
        return ""

def get_best_sellers_rank(driver):
    main_rank = ""
    sub_ranks = []
    try:
        section = driver.find_element(By.XPATH, "//span[contains(text(),'Best Sellers Rank')]/parent::span")
        try:
            main_text = section.text
            if "Best Sellers Rank:" in main_text:
                main_rank = main_text.split("Best Sellers Rank:")[1].split("\n")[0].strip()
        except:
            main_rank = ""
        try:
            items = section.find_elements(By.XPATH, ".//ul/li")
            for i in items:
                text = i.text.strip()
                if text:
                    sub_ranks.append(text)
        except:
            sub_ranks = []
    except:
        main_rank = "未找到"
        sub_ranks = ["未找到"]
    return main_rank, sub_ranks

def get_rating(driver):
    try:
        el = driver.find_element(By.CSS_SELECTOR, "span[data-hook='rating-out-of-text']")
        return el.text.strip()
    except:
        return "未找到"

def get_total_reviews(driver):
    try:
        el = driver.find_element(By.ID, "acrCustomerReviewText")
        return el.text.strip()
    except:
        return "未找到"

def get_star_percent(driver):
    stars = {}
    names = ["one", "two", "three", "four", "five"]
    for i in range(5):
        try:
            percent = driver.find_element(By.CSS_SELECTOR,f"a[href*='filterByStar={names[i]}_star'] .a-text-right").text
            stars[f"{i+1} star"] = percent
        except:
            stars[f"{i+1} star"] = "N/A"
    return stars

def get_variant_count(driver, dimension_id="flavor_name"):
    try:
        ul = driver.find_element(By.CSS_SELECTOR, f"div#inline-twister-expander-content-{dimension_id} ul.dimension-values-list")
        items = ul.find_elements(By.TAG_NAME, "li")
        if items:
            return len(items)
        ul_alt = driver.find_element(By.CSS_SELECTOR, f"div#tp-inline-twister-dim-values-container[aria-labelledby*='{dimension_id}'] ul.dimension-values-list")
        items_alt = ul_alt.find_elements(By.TAG_NAME, "li")
        return len(items_alt)
    except:
        return 0

def print_product_info(category, brand, url, title, price, coupons, discount, main_rank, sub_ranks, rating, total_reviews, stars, flavor_count, size_count):
    print("=" * 80)
    print(f"类目: {category}")
    print(f"品牌名: {brand}")
    print(f"链接: {url}")
    print(f"标题: {title}")
    print(f"大类排名: {main_rank}")
    print(f"小类排名:")
    for r in sub_ranks:
        print(f"  - {r}")
    print(f"价格: {price}")
    print(f"折扣/Deal: {discount if discount else '无'}")
    print(f"促销方式: {'; '.join(coupons) if coupons else '无'}")
    print(f"评分: {rating}")
    print(f"总评论数: {total_reviews}")
    print("星级占比:")
    for k, v in stars.items():
        print(f"  {k}: {v}")
    print(f"口味数量: {flavor_count}, 规格数量: {size_count}")
    print("=" * 80 + "\n")

def scroll_and_capture(page_name, page_index):
    screenshots = []
    last_height = driver.execute_script("return window.innerHeight")
    scroll_count = 0

    while True:
        screenshot_path = os.path.join(
            save_dir, f"{page_name}_page{page_index}_part{scroll_count+1}.png"
        )
        driver.save_screenshot(screenshot_path)
        screenshots.append(screenshot_path)

        driver.execute_script(f"window.scrollBy(0, {last_height});")
        scroll_count += 1
        time.sleep(WAIT_SCROLL)

        new_scroll_height = driver.execute_script("return window.pageYOffset + window.innerHeight")
        total_height = driver.execute_script("return document.body.scrollHeight")
        if new_scroll_height >= total_height:
            break

    return screenshots

def merge_images(image_paths, output_path):
    images = [Image.open(p) for p in image_paths]
    total_height = sum(img.height for img in images)
    max_width = max(img.width for img in images)

    merged_image = Image.new("RGB", (max_width, total_height))
    current_height = 0
    for img in images:
        merged_image.paste(img, (0, current_height))
        current_height += img.height

    merged_image.save(output_path)

    # 删除临时图
    for p in image_paths:
        os.remove(p)


def scroll_category_pages(category_name, first_page_url, max_pages=10):
    url = first_page_url
    page_index = 1
    all_screenshots = []

    while url and page_index <= max_pages:
        print(f"    📄 抓取 {category_name} 第 {page_index} 页...")
        open_amazon_page(driver, url)

        screenshots = scroll_and_capture(category_name, page_index)
        all_screenshots.extend(screenshots)

        # 翻页
        try:
            next_btn = driver.find_element(By.CSS_SELECTOR, "li.a-last a")
            if next_btn.is_enabled():
                url = next_btn.get_attribute("href")
                page_index += 1
            else:
                break
        except:
            break

    return all_screenshots


def post_to_feishu(rows, batch_size=20):
    webhook = os.getenv("FEISHU_WEBHOOK_URL", "").strip()
    if not webhook:
        print("未配置 FEISHU_WEBHOOK_URL，跳过飞书推送")
        return
    if not rows:
        print("没有成功抓取的商品数据，跳过飞书推送")
        return

    for start in range(0, len(rows), batch_size):
        batch = rows[start:start + batch_size]
        payload = {
            "source": "product_information",
            "event": "amazon_product_information_synced",
            "row_count": len(batch),
            "rows": batch,
        }
        response = requests.post(webhook, json=payload, timeout=30)
        if not response.ok:
            print(
                f"飞书推送失败: status={response.status_code}, "
                f"body={response.text[:500]}"
            )
        response.raise_for_status()
        print(f"飞书推送成功: {start + 1}-{start + len(batch)} / {len(rows)}")


db_conn = init_database(db_path)
feishu_rows = []

try:
    for file_name, sheets_dict in file_structure.items():
        print(driver.capabilities['browserVersion'])

        print(f"开始处理文件: {file_name}")

        wb = Workbook()
        wb.remove(wb.active)


        # set_amazon_zipcode(driver, "10001")

        for sheet_name, sheet_content in sheets_dict.items():
            print(f"正在抓取 Sheet: {sheet_name}")

            ws = wb.create_sheet(title=sheet_name[:31])
            ws.append([
                "类目", "品牌名", "链接", "标题", "大类排名", "小类排名", "价格",
                "折扣/Deal", "促销方式", "口味数量", "规格数量",
                "星级占比", "评分", "总评论数", "备注", "截图", "主图链接", "副图链接"
            ])

            header_fill = PatternFill(
                start_color="E2F0D9",
                end_color = "E2F0D9",
                fill_type="solid"
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            today = datetime.now().strftime("%Y-%m-%d")
            screenshot_dir = os.path.join(
                save_dir, f"Screenshots_{file_name.split('.')[0]}_{today}"
            )
            os.makedirs(screenshot_dir, exist_ok=True)


            if isinstance(sheet_content, dict) and "products" in sheet_content:
                for brand, url in sheet_content["products"].items():
                    try:
                        open_amazon_page(driver, url)
                        title = wait_title(driver)
                        screenshot_path = os.path.join(
                            screenshot_dir, f"{sheet_name}_{brand}.png"
                        )
                        driver.save_screenshot(screenshot_path)

                        price = get_price(driver)
                        coupons = get_coupon(driver)
                        discount = get_discount(driver)
                        formatted_promotion = format_promotions(discount, coupons)
                        main_rank, sub_ranks = get_best_sellers_rank(driver)
                        rating = get_rating(driver)
                        total_reviews = get_total_reviews(driver)
                        stars = get_star_percent(driver)
                        stars_text = "; ".join([f"{k}:{v}" for k, v in stars.items()])
                        flavor_count = get_variant_count(driver, "flavor_name")
                        size_count = get_variant_count(driver, "size_name")
                        image_urls = get_product_image_urls(driver)
                        main_image_url = image_urls[0] if image_urls else ""
                        sub_image_urls = image_urls[1:]

                        key = product_key(file_name, sheet_name, brand, url)
                        current_snapshot = {
                            "product_key": key,
                            "file_name": file_name,
                            "sheet_name": sheet_name,
                            "brand": brand,
                            "url": url,
                            "title": title,
                            "discount": discount if discount else "无",
                            "coupons": formatted_promotion,
                            "promotion_text": formatted_promotion,
                            "main_image_url": main_image_url,
                            "sub_image_urls": json.dumps(sub_image_urls, ensure_ascii=False),
                            "main_image_path": "",
                            "sub_image_paths": "",
                            "main_image_ocr": "",
                            "sub_image_ocr": "",
                            "price": price,
                            "stars": stars_text,
                            "rating": rating,
                            "total_reviews": total_reviews,
                            "flavor_count": str(flavor_count),
                            "size_count": str(size_count),
                            "remark": "",
                            "captured_at": datetime.now().isoformat(timespec="seconds"),
                        }
                        previous_snapshot = load_latest_snapshot(db_conn, key)
                        current_snapshot["remark"] = compare_with_previous(previous_snapshot, current_snapshot)

                        ws.append([
                            sheet_name,
                            brand,
                            url,
                            title,
                            main_rank,
                            "; ".join(sub_ranks),
                            price,
                            discount if discount else "无",
                            formatted_promotion,
                            flavor_count,
                            size_count,
                            stars_text,
                            rating,
                            total_reviews,
                            current_snapshot["remark"],
                            "",
                            main_image_url,
                            "\n".join(sub_image_urls)
                        ])

                        img = XLImage(screenshot_path)
                        img.width, img.height = 200, 150
                        ws.add_image(img, f"P{ws.max_row}")
                        ws.row_dimensions[ws.max_row].height = 120

                        save_successful_snapshot(db_conn, current_snapshot)
                        feishu_rows.append(current_snapshot)

                        print(f" {brand} 商品抓取完成")

                    except Exception as e:
                        print(f"商品抓取失败: {e}")


            if isinstance(sheet_content, dict) and "categories" in sheet_content:
                for category_name, category_url in sheet_content["categories"].items():
                    try:
                        print(f"    🌐 类目页抓取: {category_name}")

                        all_parts = scroll_category_pages(sheet_name, category_url)
                        merged_image_path = os.path.join(
                            screenshot_dir,
                            f"{sheet_name}_{category_name}_merged.png"
                        )
                        merge_images(all_parts, merged_image_path)

                        ws.append([
                            sheet_name,
                            category_name,
                            category_url,
                            "类目页截图",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            merged_image_path,
                            "",
                            ""
                        ])

                        img = XLImage(merged_image_path)
                        img.width, img.height = 300, 200
                        ws.add_image(img, f"P{ws.max_row}")
                        ws.row_dimensions[ws.max_row].height = 150

                        print("ok")

                    except Exception as e:
                        print("类目页失败")

        today_str = datetime.now().strftime("%Y-%m-%d")
        base_name = file_name.replace(".xlsx", "")
        final_file_name = f"{base_name}_{today_str}.xlsx"

        save_path = os.path.join(save_dir, final_file_name)
        wb.save(save_path)
        print(f"文件保存完成: {final_file_name}")

    post_to_feishu(feishu_rows)

finally:
    db_conn.close()
    driver.quit()
